import csv
import secrets
from datetime import datetime, timedelta

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from unicodedata import normalize

from gestion_presence.models import Cours, Etudiant, Presence, QRCode, Seance, TemporaryPresence, TemporarySeance, User
from gestion_presence.utils.qr_generator import (
    build_scan_url,
    build_qr_payload,
    make_qr_image_base64,
)


QR_VALIDITY_MINUTES = 10


def get_frontend_base_url():
    return getattr(settings, "FRONTEND_BASE_URL", "http://localhost:5173").rstrip("/")


def is_temporary_id(value):
    return str(value).startswith("temp-")


def temporary_public_id(seance):
    return f"temp-{seance.id}"


def parse_temporary_id(value):
    if not is_temporary_id(value):
        return None

    try:
        return int(str(value).split("-", 1)[1])
    except (TypeError, ValueError):
        return None


def delete_expired_temporary_seances():
    pass


def can_access_seance(user, seance):
    if user.role == User.Role.ADMIN:
        return True

    if user.role == User.Role.ENSEIGNANT:
        try:
            owner_id = seance.enseignant_id if isinstance(seance, TemporarySeance) else seance.cours.enseignant_id
            return owner_id == user.enseignant_profile.id
        except Exception:
            return False

    return False


def get_eligible_students(seance):
    module = seance.module if isinstance(seance, TemporarySeance) else seance.cours.module

    return (
        Etudiant.objects.select_related("user", "filiere")
        .filter(filiere=module.filiere)
        .distinct()
    )


def normalize_name(value):
    normalized = normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    return " ".join(normalized.lower().split())


def name_matches_student(submitted_name, etudiant):
    submitted = normalize_name(submitted_name)
    expected_full = normalize_name(f"{etudiant.user.prenom} {etudiant.user.nom}")
    expected_reverse = normalize_name(f"{etudiant.user.nom} {etudiant.user.prenom}")

    return submitted in {expected_full, expected_reverse}


def serialize_student(etudiant, presence=None):
    user = etudiant.user

    return {
        "id": etudiant.id,
        "nom": user.nom,
        "prenom": user.prenom,
        "name": f"{user.prenom} {user.nom}".strip(),
        "code_massar": etudiant.code_massar,
        "filiere": etudiant.filiere.nom,
        "validated_at": presence.heure_validation.isoformat() if presence else None,
    }


def build_temporary_qr_payload(seance, token, frontend_base_url):
    metadata = {
        "v": 1,
        "token": token,
        "seance_id": temporary_public_id(seance),
        "cours_id": temporary_public_id(seance),
        "module": seance.module.nom,
        "module_id": seance.module_id,
        "filiere": seance.module.filiere.nom,
        "filiere_id": seance.module.filiere_id,
        "niveau": "",
        "semestre": seance.module.semestre,
        "salle": seance.salle or "",
        "date": seance.date_seance.isoformat(),
        "heure": (
            f"{timezone.localtime(seance.heure_debut).strftime('%H:%M')}-"
            f"{timezone.localtime(seance.heure_fin).strftime('%H:%M')}"
        ),
        "expires_at": seance.expiration.isoformat(),
    }
    scan_url = build_scan_url(frontend_base_url, token, metadata)

    return {
        **metadata,
        "scan_url": scan_url,
    }


def serialize_qr_session(seance):
    is_temp = isinstance(seance, TemporarySeance)
    cours = None if is_temp else seance.cours
    module = seance.module if is_temp else cours.module
    qr = None if is_temp else getattr(seance, "qrcode", None)
    eligible_students = list(get_eligible_students(seance))
    presences = {
        presence.etudiant_id: presence
        for presence in seance.presences.select_related("etudiant", "etudiant__user", "etudiant__filiere").all()
    }
    attended = [
        serialize_student(etudiant, presences[etudiant.id])
        for etudiant in eligible_students
        if etudiant.id in presences
    ]
    missed = [
        serialize_student(etudiant)
        for etudiant in eligible_students
        if etudiant.id not in presences
    ]
    now = timezone.now()
    expires_at = seance.expiration if is_temp else (qr.expiration if qr else None)
    token = seance.token_qr if is_temp else (qr.token if qr else None)
    running = bool(seance.est_ouverte and token and expires_at and now < expires_at)
    frontend_base = get_frontend_base_url()
    if is_temp:
        payload = build_temporary_qr_payload(seance, token, frontend_base) if token else None
        starts_at = seance.heure_debut
        course_label = (
            f"{timezone.localtime(seance.heure_debut).strftime('%H:%M')} - "
            f"{timezone.localtime(seance.heure_fin).strftime('%H:%M')}"
        )
        room = seance.salle or "Non précisée"
        public_id = temporary_public_id(seance)
    else:
        payload = build_qr_payload(seance, token, frontend_base) if qr else None
        starts_at = seance.heure_debut
        course_label = f"{cours.heure_debut.strftime('%H:%M')} - {cours.heure_fin.strftime('%H:%M')}"
        room = cours.salle or "Non précisée"
        public_id = str(seance.id)

    now_local = timezone.localtime()
    if is_temp:
        is_still_going = seance.heure_debut <= now <= seance.heure_fin
    else:
        is_still_going = (
            seance.date_seance == now_local.date() and
            cours.heure_debut <= now_local.time() <= cours.heure_fin
        )

    return {
        "id": public_id,
        "running": running,
        "status": "active" if running else "ended",
        "date": seance.date_seance.isoformat(),
        "startsAt": starts_at.isoformat(),
        "expiresAt": expires_at.isoformat() if expires_at else None,
        "token": token,
        "scanUrl": payload["scan_url"] if payload else None,
        "qrPayload": payload,
        "qrImage": make_qr_image_base64(payload["scan_url"]) if payload else None,
        "module": module.nom,
        "filiere": module.filiere.nom,
        "niveau": "",
        "semestre": module.semestre,
        "cours": course_label,
        "room": room,
        "eligibleCount": len(eligible_students),
        "presentCount": len(attended),
        "absentCount": len(missed),
        "attended": attended,
        "missed": missed,
        "isStillGoing": is_still_going,
        "coursId": cours.id if cours else None,
        "maxDistance": seance.max_distance,
    }


def create_or_refresh_temporary_qr(seance, validite_min, latitude=None, longitude=None, max_distance=None):
    token = f"QRP-{secrets.token_urlsafe(24)}"
    expiration = timezone.now() + timedelta(minutes=validite_min)
    seance.est_ouverte = True
    seance.token_qr = token
    seance.qr_expiration = expiration
    seance.latitude = latitude
    seance.longitude = longitude
    if max_distance is not None:
        seance.max_distance = max_distance
    seance.save(update_fields=["est_ouverte", "token_qr", "qr_expiration", "latitude", "longitude", "max_distance"])
    seance.refresh_from_db()
    return seance


def create_or_refresh_qr(seance, validite_min, latitude=None, longitude=None, max_distance=None):
    token = f"QRP-{secrets.token_urlsafe(24)}"
    expiration = timezone.now() + timedelta(minutes=validite_min)
    frontend_base = get_frontend_base_url()
    payload = build_qr_payload(seance, token, frontend_base)

    seance.est_ouverte = True
    seance.token_qr = token
    seance.validite_min = validite_min
    if max_distance is not None:
        seance.max_distance = max_distance
    seance.save(update_fields=["est_ouverte", "token_qr", "validite_min", "max_distance"])

    QRCode.objects.update_or_create(
        seance=seance,
        defaults={
            "token": token,
            "url_cible": payload["scan_url"],
            "expiration": expiration,
            "latitude": latitude,
            "longitude": longitude,
        },
    )

    seance.refresh_from_db()
    return seance


@api_view(["POST"])
def generate_qr_for_cours(request):
    if request.user.role != User.Role.ENSEIGNANT:
        return Response({"message": "Seul un enseignant peut générer un QR code."}, status=status.HTTP_403_FORBIDDEN)

    cours_id = request.data.get("cours_id")
    validite_min = QR_VALIDITY_MINUTES
    latitude = request.data.get("latitude")
    longitude = request.data.get("longitude")
    max_distance = int(request.data.get("max_distance") or 20)

    try:
        enseignant = request.user.enseignant_profile
    except Exception:
        return Response({"message": "Profil enseignant introuvable."}, status=status.HTTP_403_FORBIDDEN)

    if is_temporary_id(cours_id):
        temp_id = parse_temporary_id(cours_id)
        seance = TemporarySeance.objects.select_related("module", "module__filiere").filter(
            id=temp_id,
            enseignant=enseignant,
            heure_fin__gt=timezone.now(),
        ).first()

        if not seance:
            return Response({"message": "Séance temporaire introuvable ou terminée."}, status=status.HTTP_404_NOT_FOUND)

        if seance.token_qr and timezone.now() < seance.expiration and seance.est_ouverte and not request.data.get("force"):
            return Response({"seance": serialize_qr_session(seance)})

        create_or_refresh_temporary_qr(seance, validite_min, latitude=latitude, longitude=longitude, max_distance=max_distance)
        return Response({"seance": serialize_qr_session(seance)}, status=status.HTTP_201_CREATED)

    cours = Cours.objects.select_related("module", "module__filiere").filter(
        id=cours_id,
        enseignant=enseignant,
        actif=True,
    ).first()

    if not cours:
        return Response({"message": "Cours introuvable ou non autorisé."}, status=status.HTTP_404_NOT_FOUND)

    now_local = timezone.localtime()
    if now_local.time() > cours.heure_fin:
        return Response({"message": "Ce cours est déjà terminé pour aujourd'hui."}, status=status.HTTP_400_BAD_REQUEST)

    today = now_local.date()
    heure_debut = timezone.make_aware(datetime.combine(today, cours.heure_debut))

    seance, created = Seance.objects.get_or_create(
        cours=cours,
        enseignant=enseignant,
        date_seance=today,
        defaults={
            "heure_debut": heure_debut,
            "validite_min": validite_min,
            "est_ouverte": True,
            "max_distance": max_distance,
        },
    )

    qr = getattr(seance, "qrcode", None)
    if qr and seance.est_ouverte and timezone.now() < qr.expiration and not request.data.get("force"):
        return Response({"seance": serialize_qr_session(seance)})

    if not created:
        seance.heure_debut = heure_debut
        seance.max_distance = max_distance
        seance.save(update_fields=["heure_debut", "max_distance"])

    create_or_refresh_qr(seance, validite_min, latitude=latitude, longitude=longitude, max_distance=max_distance)

    return Response({"seance": serialize_qr_session(seance)}, status=status.HTTP_201_CREATED)


@api_view(["GET"])
def current_qr_session(request):
    if request.user.role not in [User.Role.ADMIN, User.Role.ENSEIGNANT]:
        return Response({"message": "Accès non autorisé."}, status=status.HTTP_403_FORBIDDEN)

    delete_expired_temporary_seances()
    qs = (
        Seance.objects.select_related("cours", "cours__module", "cours__module__filiere", "qrcode")
        .prefetch_related("presences")
        .filter(est_ouverte=True, qrcode__expiration__gt=timezone.now())
        .order_by("-qrcode__expiration")
    )

    if request.user.role == User.Role.ENSEIGNANT:
        try:
            qs = qs.filter(cours__enseignant=request.user.enseignant_profile)
        except Exception:
            qs = Seance.objects.none()

    seance = qs.first()
    temp_qs = (
        TemporarySeance.objects.select_related("module", "module__filiere")
        .prefetch_related("presences")
        .filter(est_ouverte=True, heure_fin__gt=timezone.now(), qr_expiration__gt=timezone.now())
        .exclude(token_qr__isnull=True)
        .exclude(token_qr="")
        .order_by("-qr_expiration")
    )

    if request.user.role == User.Role.ENSEIGNANT:
        try:
            temp_qs = temp_qs.filter(enseignant=request.user.enseignant_profile)
        except Exception:
            temp_qs = TemporarySeance.objects.none()

    temp_seance = temp_qs.first()

    if temp_seance and (not seance or temp_seance.expiration > seance.qrcode.expiration):
        seance = temp_seance

    return Response({"seance": serialize_qr_session(seance) if seance else None})


@api_view(["GET"])
def qr_session_detail(request, seance_id):
    if is_temporary_id(seance_id):
        temp_id = parse_temporary_id(seance_id)
        seance = (
            TemporarySeance.objects.select_related("module", "module__filiere")
            .prefetch_related("presences")
            .filter(id=temp_id)
            .first()
        )
    else:
        seance = (
            Seance.objects.select_related("cours", "cours__module", "cours__module__filiere", "qrcode")
            .prefetch_related("presences")
            .filter(id=seance_id)
            .first()
        )

    if not seance:
        return Response({"message": "Séance introuvable."}, status=status.HTTP_404_NOT_FOUND)

    if not can_access_seance(request.user, seance):
        return Response({"message": "Accès non autorisé."}, status=status.HTTP_403_FORBIDDEN)

    return Response({"seance": serialize_qr_session(seance)})


@api_view(["POST"])
def close_qr_session(request, seance_id):
    is_temp = is_temporary_id(seance_id)
    if is_temp:
        temp_id = parse_temporary_id(seance_id)
        seance = TemporarySeance.objects.select_related("module", "module__filiere").prefetch_related("presences").filter(id=temp_id).first()
    else:
        seance = Seance.objects.select_related("cours", "cours__module", "cours__module__filiere", "qrcode").filter(id=seance_id).first()

    if not seance:
        return Response({"message": "Séance introuvable."}, status=status.HTTP_404_NOT_FOUND)

    if not can_access_seance(request.user, seance):
        return Response({"message": "Accès non autorisé."}, status=status.HTTP_403_FORBIDDEN)

    seance.est_ouverte = False
    seance.save(update_fields=["est_ouverte"])

    return Response({"seance": serialize_qr_session(seance)})


@api_view(["GET"])
def export_qr_session(request, seance_id):
    import io
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image

    if is_temporary_id(seance_id):
        temp_id = parse_temporary_id(seance_id)
        seance = TemporarySeance.objects.select_related("module", "module__filiere").prefetch_related("presences").filter(id=temp_id).first()
    else:
        seance = Seance.objects.select_related("cours", "cours__module", "cours__module__filiere", "qrcode").filter(id=seance_id).first()

    if not seance:
        return Response({"message": "Séance introuvable."}, status=status.HTTP_404_NOT_FOUND)

    if not can_access_seance(request.user, seance):
        return Response({"message": "Accès non autorisé."}, status=status.HTTP_403_FORBIDDEN)

    export_type = request.query_params.get("type", "all")
    data = serialize_qr_session(seance)
    rows = []

    if export_type in ["all", "present"]:
        rows.extend([("present", student) for student in data["attended"]])

    if export_type in ["all", "absent"]:
        rows.extend([("absent", student) for student in data["missed"]])

    # Generate workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présence"

    # Disable gridlines globally to get the clean "Word page" look
    ws.views.sheetView[0].showGridLines = False

    # Column A acts as left margin (width = 5)
    ws.column_dimensions["A"].width = 5

    # Logo resolution and insertion (anchored at B1, scaled to height = 55px)
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uni_logo.png")
    if os.path.exists(logo_path):
        img = Image(logo_path)
        scale = 55 / img.height
        img.width = int(img.width * scale)
        img.height = 55
        ws.add_image(img, "B1")

    # Header text centered across columns C to H
    ws.merge_cells("C1:H1")
    ws["C1"] = "UNIVERSITÉ MOULAY ISMAÏL"
    ws["C1"].font = Font(name="Arial", size=10, bold=True, color="1A2236")
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C2:H2")
    ws["C2"] = "FEUILLE D'ÉMARGEMENT & DE PRÉSENCE"
    ws["C2"].font = Font(name="Arial", size=12, bold=True, color="037DA7")
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C3:H3")
    ws["C3"] = f"Cours: {data.get('cours', '')} | Salle: {data.get('room', '')}"
    ws["C3"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")

    # Session details in columns B, E, and G
    filiere_nom = data.get("filiere", "")
    module_nom = data.get("module", "")
    
    is_temp = isinstance(seance, TemporarySeance)
    if is_temp:
        enseignant_name = f"{seance.enseignant.user.prenom} {seance.enseignant.user.nom}"
    else:
        enseignant_name = f"{seance.cours.enseignant.user.prenom} {seance.cours.enseignant.user.nom}"

    ws["B5"] = f"Filière : {filiere_nom}"
    ws["B5"].font = Font(name="Arial", size=9, bold=True)
    ws["E5"] = f"Module : {module_nom}"
    ws["E5"].font = Font(name="Arial", size=9, bold=True)
    ws["G5"] = f"Enseignant : {enseignant_name}"
    ws["G5"].font = Font(name="Arial", size=9, bold=True)

    # General info & summary statistics
    date_seance = data.get("date", "")
    ws["B6"] = f"Date : {date_seance} | Type d'export : {export_type.upper()}"
    ws["B6"].font = Font(name="Arial", size=9.5, bold=True, color="1A2236")
    
    ws["E6"] = f"Présents: {data['presentCount']} | Absents: {data['absentCount']} | Total: {data['eligibleCount']}"
    ws["E6"].font = Font(name="Arial", size=9.5, bold=True, color="037DA7")

    # Table Header Row at Row 8 (shifted to start from Column B (2) through H (8))
    headers = ["N°", "Statut", "Nom", "Prénom", "Code Massar", "Filière", "Heure de Validation"]
    for idx, text in enumerate(headers, 2):
        cell = ws.cell(row=8, column=idx, value=text)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A2236", end_color="1A2236", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Border styling - only applied to the data table itself to make it look like a Word grid insert
    thin_border = Border(
        left=Side(style='thin', color='DDE3EE'),
        right=Side(style='thin', color='DDE3EE'),
        top=Side(style='thin', color='DDE3EE'),
        bottom=Side(style='thin', color='DDE3EE')
    )

    # Populate table rows starting from row 9
    current_row = 9
    for idx, (row_status, student) in enumerate(rows, 1):
        ws.cell(row=current_row, column=2, value=idx) # Col B
        ws.cell(row=current_row, column=3, value="Présent" if row_status == "present" else "Absent") # Col C
        ws.cell(row=current_row, column=4, value=student["nom"]) # Col D
        ws.cell(row=current_row, column=5, value=student["prenom"]) # Col E
        ws.cell(row=current_row, column=6, value=student["code_massar"]) # Col F
        ws.cell(row=current_row, column=7, value=student["filiere"]) # Col G

        # Format validation date time
        val_time = ""
        if row_status == "present" and student["validated_at"]:
            try:
                dt = datetime.fromisoformat(student["validated_at"])
                val_time = dt.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                val_time = student["validated_at"]
        ws.cell(row=current_row, column=8, value=val_time) # Col H

        # Apply styles
        is_even = idx % 2 == 0
        row_fill = PatternFill(start_color="F8FAFC" if is_even else "FFFFFF", end_color="F8FAFC" if is_even else "FFFFFF", fill_type="solid")
        
        status_color = "16A34A" if row_status == "present" else "DC2626"
        status_font = Font(name="Arial", size=9, bold=True, color=status_color)

        for col_idx in range(2, 9): # Columns B to H
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Arial", size=9) if col_idx != 3 else status_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Center N°, Status, Massar, Validation time
            if col_idx in [2, 3, 6, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # Adjust column dimensions to content (ignoring header rows 1-6 to prevent stretching)
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col_letter == "A":
            continue
        for cell in col:
            if cell.row < 8:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Ensure validation date column has sufficient width
    ws.column_dimensions["H"].width = 24

    # Save to memory and return Excel file response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="presence_seance_{seance.id}_{export_type}.xlsx"'
    return response



@api_view(["GET"])
@permission_classes([AllowAny])
def scan_qr_info(request, token):
    qr = QRCode.objects.select_related(
        "seance",
        "seance__cours",
        "seance__cours__module",
        "seance__cours__module__filiere",
    ).filter(token=token).first()

    temp_seance = None
    if not qr:
        temp_seance = TemporarySeance.objects.select_related("module", "module__filiere").filter(token_qr=token).first()

    if not qr and not temp_seance:
        return Response({"message": "QR code introuvable."}, status=status.HTTP_404_NOT_FOUND)

    return Response({"seance": serialize_qr_session(temp_seance or qr.seance)})


import math

def calculate_distance(lat1, lon1, lat2, lon2):
    # Radius of the Earth in km
    R = 6371.0
    
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    distance_meters = R * c * 1000.0
    return distance_meters


@api_view(["POST"])
@permission_classes([AllowAny])
def submit_qr_scan(request, token):
    qr = QRCode.objects.select_related(
        "seance",
        "seance__cours",
        "seance__cours__module",
        "seance__cours__module__filiere",
    ).filter(token=token).first()

    temp_seance = None
    if not qr:
        temp_seance = TemporarySeance.objects.select_related("module", "module__filiere").filter(token_qr=token).first()

    if not qr and not temp_seance:
        return Response({"message": "QR code introuvable."}, status=status.HTTP_404_NOT_FOUND)

    seance = temp_seance or qr.seance
    is_temp = temp_seance is not None

    if (seance.est_expiree if is_temp else qr.est_expire) or not seance.est_ouverte:
        return Response({"message": "Cette séance est terminée."}, status=status.HTTP_400_BAD_REQUEST)
    code_massar = (request.data.get("code_massar") or "").strip()
    submitted_name = (request.data.get("name") or "").strip()
    if not submitted_name:
        submitted_prenom = (request.data.get("prenom") or "").strip()
        submitted_nom = (request.data.get("nom") or "").strip()
        if submitted_prenom and submitted_nom:
            submitted_name = f"{submitted_prenom} {submitted_nom}"

    if not code_massar:
        return Response({"message": "Le code Massar est obligatoire."}, status=status.HTTP_400_BAD_REQUEST)

    if not submitted_name:
        return Response({"message": "Le prénom et le nom sont obligatoires."}, status=status.HTTP_400_BAD_REQUEST)

    device_uuid = (request.data.get("device_uuid") or "").strip()

    try:
        student_lat = float(request.data.get("latitude")) if request.data.get("latitude") is not None else None
        student_lng = float(request.data.get("longitude")) if request.data.get("longitude") is not None else None
    except (ValueError, TypeError):
        student_lat = None
        student_lng = None

    teacher_lat = seance.latitude if is_temp else qr.latitude
    teacher_lng = seance.longitude if is_temp else qr.longitude

    # Enforce Geofencing if the teacher's location is set
    if teacher_lat is not None and teacher_lng is not None:
        if student_lat is None or student_lng is None:
            return Response(
                {"message": "L'autorisation de géolocalisation est obligatoire pour valider votre présence."},
                status=status.HTTP_400_BAD_REQUEST
            )

        distance = calculate_distance(teacher_lat, teacher_lng, student_lat, student_lng)
        max_distance = seance.max_distance

        if distance > max_distance:
            return Response(
                {"message": f"Vous devez être présent dans la salle de classe pour valider votre présence (vous êtes trop éloigné de l'enseignant : {distance:.1f}m, limite : {max_distance}m)."},
                status=status.HTTP_400_BAD_REQUEST
            )

    etudiant = Etudiant.objects.select_related("user", "filiere").filter(code_massar__iexact=code_massar).first()

    if not etudiant:
        return Response({"message": "Étudiant introuvable."}, status=status.HTTP_404_NOT_FOUND)

    if not name_matches_student(submitted_name, etudiant):
        return Response({"message": "Le nom ne correspond pas au code Massar."}, status=status.HTTP_403_FORBIDDEN)

    module = seance.module if is_temp else seance.cours.module
    if etudiant.filiere_id != module.filiere_id or etudiant.semestre != module.semestre:
        return Response(
            {"message": "Cet étudiant n'appartient pas à cette filière et ce semestre."},
            status=status.HTTP_403_FORBIDDEN,
        )

    if not etudiant.modules.filter(id=module.id).exists():
        return Response(
            {"message": "Cet étudiant n'est pas inscrit à ce module."},
            status=status.HTTP_403_FORBIDDEN,
        )

    presence_model = TemporaryPresence if is_temp else Presence

    # Enforce one scan per device constraint
    if device_uuid:
        existing_other = presence_model.objects.filter(
            seance=seance,
            device_uuid=device_uuid
        ).exclude(etudiant=etudiant).first()

        if existing_other:
            return Response(
                {"message": "Ce téléphone a déjà été utilisé pour valider la présence d'un autre étudiant."},
                status=status.HTTP_400_BAD_REQUEST
            )

    presence, created = presence_model.objects.get_or_create(
        etudiant=etudiant,
        seance=seance,
        defaults={"statut": presence_model.Statut.PRESENT, "device_uuid": device_uuid},
    )

    if not created:
        update_fields = []
        if presence.statut != presence_model.Statut.PRESENT:
            presence.statut = presence_model.Statut.PRESENT
            update_fields.append("statut")
        if device_uuid and presence.device_uuid != device_uuid:
            presence.device_uuid = device_uuid
            update_fields.append("device_uuid")
        if update_fields:
            presence.save(update_fields=update_fields)

    return Response(
        {
            "message": "Presence enregistree." if created else "Presence deja enregistree.",
            "student": serialize_student(etudiant, presence),
            "already_registered": not created,
        }
    )
