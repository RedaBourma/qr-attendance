from django.db import IntegrityError, transaction
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from gestion_presence.models import Enseignant, User
from gestion_presence.api.auth_views import serialize_user, file_to_base64_data_uri


def serialize_enseignant(enseignant, request=None):
    user = enseignant.user

    return {
        "id": enseignant.id,
        "user": serialize_user(user, request),
        "cours_count": enseignant.cours.count(),
        "filieres": [{"id": f.id, "nom": f.nom} for f in enseignant.filieres.all()],
        "modules": [{"id": m.id, "nom": m.nom, "filiere_id": m.filiere_id, "semestre": m.semestre} for m in enseignant.modules.all()],
    }


@api_view(["GET"])
def list_enseignants(request):
    if request.user.role != User.Role.ADMIN:
        return Response(
            {"message": "Accès réservé à l'admin."},
            status=status.HTTP_403_FORBIDDEN,
        )

    enseignants = Enseignant.objects.select_related("user").prefetch_related("cours", "filieres", "modules").all()
    return Response({"enseignants": [serialize_enseignant(enseignant, request) for enseignant in enseignants]})


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def create_enseignant(request):
    if request.user.role != User.Role.ADMIN:
        return Response(
            {"message": "Accès réservé à l'admin."},
            status=status.HTTP_403_FORBIDDEN,
        )

    nom = (request.data.get("nom") or "").strip()
    prenom = (request.data.get("prenom") or "").strip()
    email = (request.data.get("email") or "").strip()
    password = request.data.get("password") or ""

    if not all([nom, prenom]):
        return Response(
            {"message": "nom et prenom sont requis."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    clean_nom = nom.replace(" ", "").lower()
    clean_prenom = prenom.replace(" ", "").lower()

    if not email:
        email = f"{clean_prenom[0]}.{clean_nom}@umi.ac.ma"

    if not password:
        password = f"{clean_nom}@{clean_prenom}"

    # Handle lists from either JSON or multipart form data
    filiere_ids = request.data.getlist("filiere_ids")
    if not filiere_ids and "filiere_ids" in request.data:
        val = request.data.get("filiere_ids")
        if isinstance(val, list):
            filiere_ids = val
        elif isinstance(val, str):
            import json
            try:
                filiere_ids = json.loads(val)
            except ValueError:
                pass

    module_ids = request.data.getlist("module_ids")
    if not module_ids and "module_ids" in request.data:
        val = request.data.get("module_ids")
        if isinstance(val, list):
            module_ids = val
        elif isinstance(val, str):
            import json
            try:
                module_ids = json.loads(val)
            except ValueError:
                pass

    profile_picture = request.FILES.get("profile_picture")

    try:
        with transaction.atomic():
            user = User.objects.create_user(
                email=email,
                password=password,
                nom=nom,
                prenom=prenom,
                role=User.Role.ENSEIGNANT,
            )
            if profile_picture:
                user.profile_picture = file_to_base64_data_uri(profile_picture)
                user.save(update_fields=["profile_picture"])

            enseignant = Enseignant.objects.create(user=user)
            if filiere_ids:
                enseignant.filieres.set(filiere_ids)
            if module_ids:
                enseignant.modules.set(module_ids)
    except IntegrityError:
        return Response(
            {"message": "Un utilisateur avec cet email existe deja."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    return Response(
        serialize_enseignant(enseignant, request),
        status=status.HTTP_201_CREATED,
    )


@api_view(["PATCH", "POST"])
@parser_classes([MultiPartParser, FormParser])
def update_enseignant(request, enseignant_id):
    if request.user.role != User.Role.ADMIN:
        return Response(
            {"message": "Accès réservé à l'admin."},
            status=status.HTTP_403_FORBIDDEN,
        )

    enseignant = Enseignant.objects.filter(id=enseignant_id).select_related("user").first()
    if not enseignant:
        return Response({"message": "Enseignant introuvable."}, status=status.HTTP_404_NOT_FOUND)

    nom = (request.data.get("nom") or "").strip()
    prenom = (request.data.get("prenom") or "").strip()
    email = (request.data.get("email") or "").strip()
    password = request.data.get("password") or ""

    if not all([nom, prenom]):
        return Response(
            {"message": "nom et prenom sont requis."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Handle lists from either JSON or multipart form data
    filiere_ids = request.data.getlist("filiere_ids")
    if not filiere_ids and "filiere_ids" in request.data:
        val = request.data.get("filiere_ids")
        if isinstance(val, list):
            filiere_ids = val
        elif isinstance(val, str):
            import json
            try:
                filiere_ids = json.loads(val)
            except ValueError:
                pass

    module_ids = request.data.getlist("module_ids")
    if not module_ids and "module_ids" in request.data:
        val = request.data.get("module_ids")
        if isinstance(val, list):
            module_ids = val
        elif isinstance(val, str):
            import json
            try:
                module_ids = json.loads(val)
            except ValueError:
                pass

    profile_picture = request.FILES.get("profile_picture")
    remove_picture = request.data.get("remove_profile_picture") == "true" or request.data.get("remove_profile_picture") is True

    try:
        with transaction.atomic():
            user = enseignant.user
            user.nom = nom
            user.prenom = prenom
            if email:
                user.email = email
            if password:
                user.set_password(password)

            if remove_picture:
                user.profile_picture = None
            elif profile_picture:
                user.profile_picture = file_to_base64_data_uri(profile_picture)

            user.save()

            enseignant.filieres.set(filiere_ids)
            enseignant.modules.set(module_ids)
    except IntegrityError:
        return Response(
            {"message": "Un utilisateur avec cet email existe deja."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    return Response(
        serialize_enseignant(enseignant, request),
        status=status.HTTP_200_OK,
    )


@api_view(["DELETE", "POST"])
def delete_enseignant(request, enseignant_id):
    if request.user.role != User.Role.ADMIN:
        return Response(
            {"message": "Accès réservé à l'admin."},
            status=status.HTTP_403_FORBIDDEN,
        )

    enseignant = Enseignant.objects.filter(id=enseignant_id).select_related("user").first()
    if not enseignant:
        return Response({"message": "Enseignant introuvable."}, status=status.HTTP_404_NOT_FOUND)

    try:
        with transaction.atomic():
            user = enseignant.user
            user.delete()
    except Exception as e:
        return Response(
            {"message": f"Erreur lors de la suppression : {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    return Response({"message": "Enseignant supprimé avec succès."}, status=status.HTTP_200_OK)


@api_view(["POST"])
def import_enseignants(request):
    if request.user.role != User.Role.ADMIN:
        return Response(
            {"message": "Accès réservé à l'admin."},
            status=status.HTTP_403_FORBIDDEN,
        )

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"message": "Aucun fichier n'a été fourni."}, status=status.HTTP_400_BAD_REQUEST)

    import openpyxl
    import csv
    from io import TextIOWrapper

    file_name = uploaded_file.name.lower()
    rows = []

    if file_name.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            sheet = wb.active
            if not sheet:
                return Response({"message": "Le fichier Excel est vide ou invalide."}, status=status.HTTP_400_BAD_REQUEST)
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  
                    continue
                prenom = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                nom = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                email = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                if prenom or nom:
                    rows.append({
                        "prenom": prenom,
                        "nom": nom,
                        "email": email
                    })
        except Exception as e:
            return Response({"message": f"Erreur lors de la lecture du fichier Excel : {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
    elif file_name.endswith(".csv"):
        try:
            csv_file = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
            sample = csv_file.read(2048)
            csv_file.seek(0)
            delimiter = ";" if ";" in sample else ","
            reader = csv.reader(csv_file, delimiter=delimiter)
            header = next(reader, None)  
            for row in reader:
                if not any(row):
                    continue
                prenom = row[0].strip() if len(row) > 0 else ""
                nom = row[1].strip() if len(row) > 1 else ""
                email = row[2].strip() if len(row) > 2 else ""
                if prenom or nom:
                    rows.append({
                        "prenom": prenom,
                        "nom": nom,
                        "email": email
                    })
        except Exception as e:
            return Response({"message": f"Erreur lors de la lecture du fichier CSV : {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"message": "Le format du fichier doit être Excel (.xlsx) ou CSV (.csv)."}, status=status.HTTP_400_BAD_REQUEST)

    if not rows:
        return Response({"message": "Aucun enseignant valide n'a été trouvé dans le fichier. Veuillez vous assurer que le fichier contient des données à partir de la deuxième ligne."}, status=status.HTTP_400_BAD_REQUEST)

    success_count = 0
    errors = []

    for index, row_data in enumerate(rows, start=2):
        prenom = row_data["prenom"]
        nom = row_data["nom"]
        email = row_data["email"]

        if not prenom or not nom:
            errors.append(f"Ligne {index} : Le prénom et le nom sont obligatoires.")
            continue

        clean_nom = nom.replace(" ", "").lower()
        clean_prenom = prenom.replace(" ", "").lower()

        if not email:
            email = f"{clean_prenom[0]}.{clean_nom}@umi.ac.ma"

        password = f"{clean_nom}@{clean_prenom}"

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    nom=nom,
                    prenom=prenom,
                    role=User.Role.ENSEIGNANT,
                )
                Enseignant.objects.create(user=user)
                success_count += 1
        except IntegrityError:
            errors.append(f"Ligne {index} ({prenom} {nom}) : Un utilisateur avec cet email existe déjà.")
        except Exception as e:
            errors.append(f"Ligne {index} ({prenom} {nom}) : {str(e)}")

    return Response({
        "success": success_count,
        "total": len(rows),
        "errors": errors
    }, status=status.HTTP_200_OK if success_count > 0 else status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def download_teacher_template(request):
    if request.user.role != User.Role.ADMIN:
        return Response({"message": "Accès réservé à l'admin."}, status=status.HTTP_403_FORBIDDEN)

    from django.http import HttpResponse
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Enseignants"
    ws.append(["Prénom", "Nom", "Email"])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="template_enseignants.xlsx"'
    wb.save(response)
    return response
