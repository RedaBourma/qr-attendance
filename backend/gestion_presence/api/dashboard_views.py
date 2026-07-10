from datetime import timedelta

from django.db.models import Count
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from gestion_presence.models import Cours, Enseignant, Etudiant, Filiere, Module, Presence, Seance, TemporaryPresence, TemporarySeance, User


def get_visible_seances(user):
    seances = (
        Seance.objects.select_related("cours", "cours__module", "cours__module__filiere", "cours__enseignant")
        .prefetch_related("presences")
        .annotate(presence_count=Count("presences"))
        .filter(qrcode__isnull=False)
    )

    if user.role == User.Role.ADMIN:
        return seances

    if user.role == User.Role.ENSEIGNANT:
        try:
            return seances.filter(cours__enseignant=user.enseignant_profile)
        except Exception:
            return Seance.objects.none()

    return Seance.objects.none()


def get_visible_temporary_seances(user):
    seances = (
        TemporarySeance.objects.select_related(
            "module",
            "module__filiere",
            "enseignant",
            "enseignant__user",
        )
        .prefetch_related("presences")
        .annotate(presence_count=Count("presences"))
    )

    if user.role == User.Role.ADMIN:
        return seances

    if user.role == User.Role.ENSEIGNANT:
        try:
            return seances.filter(enseignant=user.enseignant_profile)
        except Exception:
            return TemporarySeance.objects.none()

    return TemporarySeance.objects.none()


def get_expiration(seance):
    try:
        return seance.qrcode.expiration
    except Exception:
        return seance.heure_debut + timedelta(minutes=seance.validite_min)


def serialize_seance(seance):
    expiration = get_expiration(seance)
    is_active = seance.est_ouverte and timezone.now() < expiration
    cours = seance.cours

    return {
        "id": str(seance.id),
        "module": cours.module.nom,
        "filiere": cours.module.filiere.nom,
        "semestre": cours.module.semestre,
        "cours": f"{cours.heure_debut.strftime('%H:%M')} - {cours.heure_fin.strftime('%H:%M')} | Salle {cours.salle or 'Non precisee'}",
        "date": seance.date_seance.isoformat(),
        "startsAt": seance.heure_debut.isoformat(),
        "expiresAt": expiration.isoformat(),
        "status": "active" if is_active else "expired",
        "presences": seance.presence_count,
        "scanUrl": seance.qrcode.url_cible if hasattr(seance, "qrcode") else None,
        "isTemporary": False,
        "enseignant": str(cours.enseignant),
        "professor": None,
    }


def serialize_temporary_seance(seance):
    expiration = seance.expiration
    is_active = seance.est_ouverte and timezone.now() < expiration
    module = seance.module
    professor_user = seance.enseignant.user

    return {
        "id": f"temp-{seance.id}",
        "module": module.nom,
        "filiere": module.filiere.nom,
        "semestre": module.semestre,
        "cours": f"{timezone.localtime(seance.heure_debut).strftime('%H:%M')} - {timezone.localtime(seance.heure_fin).strftime('%H:%M')} | Salle {seance.salle or 'Non precisee'}",
        "date": seance.date_seance.isoformat(),
        "startsAt": seance.heure_debut.isoformat(),
        "expiresAt": expiration.isoformat(),
        "status": "active" if is_active else "expired",
        "presences": seance.presence_count,
        "scanUrl": None,
        "isTemporary": True,
        "enseignant": str(seance.enseignant),
        "professor": {
            "id": seance.enseignant_id,
            "nom": professor_user.nom,
            "prenom": professor_user.prenom,
            "email": professor_user.email,
        },
    }


def rate(present, eligible):
    return round((present / eligible) * 100) if eligible else 0


def presence_counts_by_student(model):
    return {
        item["etudiant_id"]: item["total"]
        for item in model.objects.values("etudiant_id").annotate(total=Count("id"))
    }


def build_analytics(seances, temporary_seances):
    students = list(Etudiant.objects.select_related("user", "filiere").all())
    teachers = list(Enseignant.objects.select_related("user").all())
    courses = list(Cours.objects.select_related("module", "module__filiere", "enseignant", "enseignant__user").filter(actif=True))

    eligible_by_module = {}
    for item in Etudiant.objects.values("filiere_id").annotate(total=Count("id")).order_by():
        eligible_by_module[item["filiere_id"]] = item["total"]

    session_count_by_scope = {}
    total_eligible = 0

    course_rows = {
        cours.id: {
            "id": cours.id,
            "module": cours.module.nom,
            "filiere": cours.module.filiere.nom,
            "semestre": cours.module.semestre,
            "enseignant": str(cours.enseignant),
            "sessions": 0,
            "presences": 0,
            "eligible": 0,
            "attendanceRate": 0,
        }
        for cours in courses
    }

    teacher_rows = {
        teacher.id: {
            "id": teacher.id,
            "name": str(teacher),
            "email": teacher.user.email,
            "courses": 0,
            "sessions": 0,
            "temporarySessions": 0,
            "presences": 0,
            "eligible": 0,
            "attendanceRate": 0,
        }
        for teacher in teachers
    }

    for cours in courses:
        if cours.enseignant_id in teacher_rows:
            teacher_rows[cours.enseignant_id]["courses"] += 1

    for seance in seances:
        module = seance.cours.module
        filiere_id = module.filiere_id
        eligible = eligible_by_module.get(filiere_id, 0)
        presences = seance.presence_count
        total_eligible += eligible
        session_count_by_scope[filiere_id] = session_count_by_scope.get(filiere_id, 0) + 1

        if seance.cours_id in course_rows:
            course_rows[seance.cours_id]["sessions"] += 1
            course_rows[seance.cours_id]["presences"] += presences
            course_rows[seance.cours_id]["eligible"] += eligible

        if seance.enseignant_id in teacher_rows:
            teacher_rows[seance.enseignant_id]["sessions"] += 1
            teacher_rows[seance.enseignant_id]["presences"] += presences
            teacher_rows[seance.enseignant_id]["eligible"] += eligible

    for seance in temporary_seances:
        module = seance.module
        filiere_id = module.filiere_id
        eligible = eligible_by_module.get(filiere_id, 0)
        presences = seance.presence_count
        total_eligible += eligible
        session_count_by_scope[filiere_id] = session_count_by_scope.get(filiere_id, 0) + 1

        if seance.enseignant_id in teacher_rows:
            teacher_rows[seance.enseignant_id]["sessions"] += 1
            teacher_rows[seance.enseignant_id]["temporarySessions"] += 1
            teacher_rows[seance.enseignant_id]["presences"] += presences
            teacher_rows[seance.enseignant_id]["eligible"] += eligible

    regular_presence_counts = presence_counts_by_student(Presence)
    temporary_presence_counts = presence_counts_by_student(TemporaryPresence)
    student_rows = []

    for student in students:
        eligible = session_count_by_scope.get(student.filiere_id, 0)
        present = regular_presence_counts.get(student.id, 0) + temporary_presence_counts.get(student.id, 0)
        student_rows.append(
            {
                "id": student.id,
                "name": f"{student.user.prenom} {student.user.nom}".strip(),
                "email": student.user.email,
                "codeMassar": student.code_massar,
                "filiere": student.filiere.nom,
                "eligibleSessions": eligible,
                "presences": present,
                "absences": max(eligible - present, 0),
                "attendanceRate": rate(present, eligible),
            }
        )

    for row in course_rows.values():
        row["attendanceRate"] = rate(row["presences"], row["eligible"])

    for row in teacher_rows.values():
        row["attendanceRate"] = rate(row["presences"], row["eligible"])

    return {
        "summary": {
            "totalStudents": len(students),
            "totalTeachers": len(teachers),
            "totalCourses": len(courses),
            "totalEligible": total_eligible,
        },
        "students": sorted(student_rows, key=lambda item: (item["attendanceRate"], -item["eligibleSessions"], item["name"])),
        "courses": sorted(course_rows.values(), key=lambda item: (-item["attendanceRate"], item["module"])),
        "teachers": sorted(teacher_rows.values(), key=lambda item: (-item["attendanceRate"], item["name"])),
    }


@api_view(["GET"])
def dashboard_view(request):
    user = request.user

    if user.role not in (User.Role.ADMIN, User.Role.ENSEIGNANT):
        return Response(
            {"message": "Acces non autorise."},
            status=status.HTTP_403_FORBIDDEN,
        )

    seances = list(get_visible_seances(user))
    temporary_seances = list(get_visible_temporary_seances(user))
    serialized_all = [serialize_seance(seance) for seance in seances] + [
        serialize_temporary_seance(seance) for seance in temporary_seances
    ]
    serialized_all.sort(key=lambda seance: seance["startsAt"], reverse=True)
    total_seances = len(serialized_all)
    active_count = sum(1 for seance in serialized_all if seance["status"] == "active")
    total_presences = sum(seance["presences"] for seance in serialized_all)
    avg_presences = round(total_presences / total_seances) if total_seances else 0

    if user.role == User.Role.ENSEIGNANT:
        return Response(
            {
                "stats": {
                    "totalSeances": total_seances,
                    "activeSeances": active_count,
                    "totalPresences": total_presences,
                    "avgPresences": avg_presences,
                },
                "seances": serialized_all,
            }
        )

    analytics = build_analytics(seances, temporary_seances)

    return Response(
        {
            "stats": {
                "totalSeances": total_seances,
                "activeSeances": active_count,
                "totalPresences": total_presences,
                "avgPresences": avg_presences,
                "attendanceRate": rate(total_presences, analytics["summary"]["totalEligible"]),
                **analytics["summary"],
            },
            "analytics": {
                "students": analytics["students"],
                "courses": analytics["courses"],
                "teachers": analytics["teachers"],
            },
            "seances": serialized_all[:20],
        }
    )


@api_view(["GET"])
def custom_export_view(request):
    import io
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime
    from django.db.models import Count

    user = request.user
    if user.role != User.Role.ADMIN:
        return Response({"message": "Accès non autorisé."}, status=status.HTTP_403_FORBIDDEN)

    scope = request.query_params.get("scope")
    export_type = request.query_params.get("type", "all")
    semestre_filter = request.query_params.get("semestre")
    if semestre_filter == "Tous":
        semestre_filter = None

    def apply_common_styles(ws, title, subtitle_text):
        ws.views.sheetView[0].showGridLines = False
        ws.column_dimensions["A"].width = 5

        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uni_logo.png")
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                scale = 55 / img.height
                img.width = int(img.width * scale)
                img.height = 55
                ws.add_image(img, "B1")
            except Exception:
                pass

        ws.merge_cells("C1:H1")
        ws["C1"] = "UNIVERSITÉ MOULAY ISMAÏL"
        ws["C1"].font = Font(name="Arial", size=10, bold=True, color="1A2236")
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("C2:H2")
        ws["C2"] = title
        ws["C2"].font = Font(name="Arial", size=12, bold=True, color="037DA7")
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("C3:H3")
        ws["C3"] = subtitle_text
        ws["C3"].font = Font(name="Arial", size=9, italic=True, color="64748B")
        ws["C3"].alignment = Alignment(horizontal="center", vertical="center")

    def write_table_headers(ws, start_row, headers):
        for idx, text in enumerate(headers, 2):
            cell = ws.cell(row=start_row, column=idx, value=text)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A2236", end_color="1A2236", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_table_rows(ws, start_row, row_count, status_col_idx=None, center_cols=None):
        thin_border = Border(
            left=Side(style='thin', color='DDE3EE'),
            right=Side(style='thin', color='DDE3EE'),
            top=Side(style='thin', color='DDE3EE'),
            bottom=Side(style='thin', color='DDE3EE')
        )

        for r_idx in range(start_row, start_row + row_count):
            is_even = (r_idx - start_row) % 2 == 1
            row_fill = PatternFill(start_color="F8FAFC" if is_even else "FFFFFF", end_color="F8FAFC" if is_even else "FFFFFF", fill_type="solid")

            max_col = ws.max_column
            for col_idx in range(2, max_col + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = Font(name="Arial", size=9)
                cell.fill = row_fill
                cell.border = thin_border

                if center_cols and col_idx in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if status_col_idx and col_idx == status_col_idx:
                    val = str(cell.value or '').lower()
                    if val in ["présent", "present", "retard"]:
                        cell.font = Font(name="Arial", size=9, bold=True, color="16A34A")
                    elif val in ["absent"]:
                        cell.font = Font(name="Arial", size=9, bold=True, color="DC2626")

    def auto_fit_columns(ws, start_row=8):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            if col_letter == "A":
                continue
            for cell in col:
                if cell.row < start_row:
                    continue
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb = openpyxl.Workbook()
    filename = "export_custom.xlsx"

    if scope == "student":
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response({"message": "ID étudiant requis."}, status=status.HTTP_400_BAD_REQUEST)
        student = Etudiant.objects.select_related("user", "filiere").filter(id=student_id).first()
        if not student:
            return Response({"message": "Étudiant introuvable."}, status=status.HTTP_404_NOT_FOUND)

        modules = Module.objects.filter(filiere=student.filiere)
        if semestre_filter:
            modules = modules.filter(semestre=semestre_filter)
        module_ids = list(modules.values_list("id", flat=True))

        seances = Seance.objects.filter(cours__module_id__in=module_ids, qrcode__isnull=False).select_related(
            "cours", "cours__module", "cours__enseignant", "cours__enseignant__user"
        )
        temp_seances = TemporarySeance.objects.filter(module_id__in=module_ids).select_related(
            "module", "enseignant", "enseignant__user"
        )

        presences = {p.seance_id: p for p in Presence.objects.filter(etudiant=student, seance__in=seances)}
        temp_presences = {p.seance_id: p for p in TemporaryPresence.objects.filter(etudiant=student, seance__in=temp_seances)}

        rows = []
        for s in seances:
            p = presences.get(s.id)
            stat_val = p.statut if p else "absent"
            val_time = p.heure_validation if (p and p.heure_validation) else None
            rows.append({
                "date": s.date_seance,
                "semestre": s.cours.module.semestre,
                "module": s.cours.module.nom,
                "enseignant": str(s.cours.enseignant),
                "salle": s.cours.salle or "Non précisée",
                "status": stat_val,
                "validated_at": val_time
            })

        for ts in temp_seances:
            p = temp_presences.get(ts.id)
            stat_val = p.statut if p else "absent"
            val_time = p.heure_validation if (p and p.heure_validation) else None
            rows.append({
                "date": ts.date_seance,
                "semestre": ts.module.semestre,
                "module": ts.module.nom,
                "enseignant": str(ts.enseignant),
                "salle": ts.salle or "Non précisée",
                "status": stat_val,
                "validated_at": val_time
            })

        rows.sort(key=lambda r: r["date"], reverse=True)

        if export_type == "present":
            rows = [r for r in rows if r["status"] in ["present", "retard"]]
        elif export_type == "absent":
            rows = [r for r in rows if r["status"] == "absent"]

        ws = wb.active
        ws.title = "Fiche Étudiant"
        
        apply_common_styles(ws, "FICHE DE PRÉSENCE ÉTUDIANT", f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        ws["B5"] = f"Étudiant : {student.user.prenom} {student.user.nom}"
        ws["B5"].font = Font(name="Arial", size=9, bold=True)
        ws["E5"] = f"Code Massar : {student.code_massar}"
        ws["E5"].font = Font(name="Arial", size=9, bold=True)
        ws["G5"] = f"Filière : {student.filiere.nom}"
        ws["G5"].font = Font(name="Arial", size=9, bold=True)

        pres_cnt = sum(1 for r in rows if r["status"] in ["present", "retard"])
        abs_cnt = sum(1 for r in rows if r["status"] == "absent")
        tot_cnt = len(rows)
        rate_val = round((pres_cnt / tot_cnt) * 100) if tot_cnt else 0

        ws["B6"] = f"Période : {semestre_filter or 'Tous les semestres'}"
        ws["B6"].font = Font(name="Arial", size=9.5, bold=True)
        ws["E6"] = f"Présents: {pres_cnt} | Absents: {abs_cnt} | Taux: {rate_val}%"
        ws["E6"].font = Font(name="Arial", size=9.5, bold=True, color="037DA7")

        headers = ["N°", "Date", "Semestre", "Module", "Enseignant", "Salle", "Statut", "Heure de Validation"]
        write_table_headers(ws, 8, headers)

        current_row = 9
        for idx, r in enumerate(rows, 1):
            ws.cell(row=current_row, column=2, value=idx)
            ws.cell(row=current_row, column=3, value=r["date"].strftime("%d/%m/%Y") if isinstance(r["date"], datetime) or hasattr(r["date"], "strftime") else str(r["date"]))
            ws.cell(row=current_row, column=4, value=r["semestre"])
            ws.cell(row=current_row, column=5, value=r["module"])
            ws.cell(row=current_row, column=6, value=r["enseignant"])
            ws.cell(row=current_row, column=7, value=r["salle"])
            ws.cell(row=current_row, column=8, value="Présent" if r["status"] in ["present", "retard"] else "Absent")

            val_time_str = ""
            if r["validated_at"]:
                try:
                    val_time_str = timezone.localtime(r["validated_at"]).strftime("%d/%m/%Y %H:%M:%S")
                except Exception:
                    val_time_str = str(r["validated_at"])
            ws.cell(row=current_row, column=9, value=val_time_str)
            current_row += 1

        style_table_rows(ws, 9, len(rows), status_col_idx=8, center_cols=[2, 3, 4, 7, 8, 9])
        auto_fit_columns(ws, start_row=8)
        ws.column_dimensions["I"].width = 24
        
        filename = f"presence_etudiant_{student.code_massar}_{export_type}.xlsx"

    elif scope == "class":
        filiere_id = request.query_params.get("filiere_id")
        if not filiere_id:
            return Response({"message": "ID classe/filière requis."}, status=status.HTTP_400_BAD_REQUEST)
        filiere = Filiere.objects.filter(id=filiere_id).first()
        if not filiere:
            return Response({"message": "Classe/Filière introuvable."}, status=status.HTTP_404_NOT_FOUND)

        students = Etudiant.objects.filter(filiere=filiere).select_related("user")
        modules = Module.objects.filter(filiere=filiere)
        if semestre_filter:
            modules = modules.filter(semestre=semestre_filter)
        module_ids = list(modules.values_list("id", flat=True))

        seances = Seance.objects.filter(cours__module_id__in=module_ids, qrcode__isnull=False).select_related(
            "cours", "cours__module"
        )
        temp_seances = TemporarySeance.objects.filter(module_id__in=module_ids).select_related("module")

        total_sessions = len(seances) + len(temp_seances)

        presence_by_student_seance = {}
        presence_times = {}

        for p in Presence.objects.filter(seance_id__in=seances):
            presence_by_student_seance[(p.etudiant_id, str(p.seance_id))] = p.statut
            presence_times[(p.etudiant_id, str(p.seance_id))] = p.heure_validation

        for p in TemporaryPresence.objects.filter(seance_id__in=temp_seances):
            presence_by_student_seance[(p.etudiant_id, f"temp-{p.seance_id}")] = p.statut
            presence_times[(p.etudiant_id, f"temp-{p.seance_id}")] = p.heure_validation

        synthesis = []
        for std in students:
            present_count = 0
            absent_count = 0

            for s in seances:
                stat = presence_by_student_seance.get((std.id, str(s.id)), "absent")
                if stat in ["present", "retard"]:
                    present_count += 1
                else:
                    absent_count += 1

            for ts in temp_seances:
                stat = presence_by_student_seance.get((std.id, f"temp-{ts.id}"), "absent")
                if stat in ["present", "retard"]:
                    present_count += 1
                else:
                    absent_count += 1

            rate_val = round((present_count / total_sessions) * 100) if total_sessions else 0
            synthesis.append({
                "student": std,
                "eligible": total_sessions,
                "present": present_count,
                "absent": absent_count,
                "rate": rate_val
            })

        details = []
        for std in students:
            for s in seances:
                stat = presence_by_student_seance.get((std.id, str(s.id)), "absent")
                if export_type == "present" and stat == "absent":
                    continue
                if export_type == "absent" and stat in ["present", "retard"]:
                    continue
                val_time = presence_times.get((std.id, str(s.id)))
                details.append({
                    "date": s.date_seance,
                    "module": s.cours.module.nom,
                    "semestre": s.cours.module.semestre,
                    "student_name": f"{std.user.prenom} {std.user.nom}",
                    "code_massar": std.code_massar,
                    "status": stat,
                    "validated_at": val_time
                })
            for ts in temp_seances:
                stat = presence_by_student_seance.get((std.id, f"temp-{ts.id}"), "absent")
                if export_type == "present" and stat == "absent":
                    continue
                if export_type == "absent" and stat in ["present", "retard"]:
                    continue
                val_time = presence_times.get((std.id, f"temp-{ts.id}"))
                details.append({
                    "date": ts.date_seance,
                    "module": ts.module.nom,
                    "semestre": ts.module.semestre,
                    "student_name": f"{std.user.prenom} {std.user.nom}",
                    "code_massar": std.code_massar,
                    "status": stat,
                    "validated_at": val_time
                })

        details.sort(key=lambda d: d["date"], reverse=True)

        ws1 = wb.active
        ws1.title = "Synthèse Classe"
        apply_common_styles(ws1, "SYNTHÈSE DE PRÉSENCE CLASSE", f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        ws1["B5"] = f"Classe / Filière : {filiere.nom}"
        ws1["B5"].font = Font(name="Arial", size=9, bold=True)
        ws1["E5"] = f"Semestre : {semestre_filter or 'Tous les semestres'}"
        ws1["E5"].font = Font(name="Arial", size=9, bold=True)
        ws1["G5"] = f"Effectif : {len(students)} étudiants"
        ws1["G5"].font = Font(name="Arial", size=9, bold=True)

        ws1["B6"] = f"Séances totales : {total_sessions}"
        ws1["B6"].font = Font(name="Arial", size=9.5, bold=True, color="037DA7")

        headers1 = ["N°", "Code Massar", "Nom", "Prénom", "Séances Éligibles", "Présences", "Absences", "Taux de Présence"]
        write_table_headers(ws1, 8, headers1)

        current_row = 9
        for idx, row in enumerate(synthesis, 1):
            ws1.cell(row=current_row, column=2, value=idx)
            ws1.cell(row=current_row, column=3, value=row["student"].code_massar)
            ws1.cell(row=current_row, column=4, value=row["student"].user.nom)
            ws1.cell(row=current_row, column=5, value=row["student"].user.prenom)
            ws1.cell(row=current_row, column=6, value=row["eligible"])
            ws1.cell(row=current_row, column=7, value=row["present"])
            ws1.cell(row=current_row, column=8, value=row["absent"])
            ws1.cell(row=current_row, column=9, value=f"{row['rate']}%")
            current_row += 1

        style_table_rows(ws1, 9, len(synthesis), center_cols=[2, 3, 6, 7, 8, 9])
        auto_fit_columns(ws1, start_row=8)

        ws2 = wb.create_sheet(title="Détail Émargements")
        apply_common_styles(ws2, "DÉTAIL DES ÉMARGEMENTS CLASSE", f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        ws2["B5"] = f"Classe / Filière : {filiere.nom}"
        ws2["B5"].font = Font(name="Arial", size=9, bold=True)
        ws2["E5"] = f"Type d'export : {export_type.upper()}"
        ws2["E5"].font = Font(name="Arial", size=9, bold=True)

        headers2 = ["N°", "Date", "Module", "Semestre", "Nom", "Prénom", "Code Massar", "Statut", "Heure de Validation"]
        write_table_headers(ws2, 8, headers2)

        current_row = 9
        for idx, r in enumerate(details, 1):
            ws2.cell(row=current_row, column=2, value=idx)
            ws2.cell(row=current_row, column=3, value=r["date"].strftime("%d/%m/%Y") if isinstance(r["date"], datetime) or hasattr(r["date"], "strftime") else str(r["date"]))
            ws2.cell(row=current_row, column=4, value=r["module"])
            ws2.cell(row=current_row, column=5, value=r["semestre"])
            names = r["student_name"].split(" ")
            first_name = names[0]
            last_name = " ".join(names[1:]) if len(names) > 1 else ""
            ws2.cell(row=current_row, column=6, value=last_name)
            ws2.cell(row=current_row, column=7, value=first_name)
            ws2.cell(row=current_row, column=8, value=r["code_massar"])
            ws2.cell(row=current_row, column=9, value="Présent" if r["status"] in ["present", "retard"] else "Absent")

            val_time_str = ""
            if r["validated_at"]:
                try:
                    val_time_str = timezone.localtime(r["validated_at"]).strftime("%d/%m/%Y %H:%M:%S")
                except Exception:
                    val_time_str = str(r["validated_at"])
            ws2.cell(row=current_row, column=10, value=val_time_str)
            current_row += 1

        style_table_rows(ws2, 9, len(details), status_col_idx=9, center_cols=[2, 3, 5, 8, 9, 10])
        auto_fit_columns(ws2, start_row=8)
        ws2.column_dimensions["J"].width = 24

        filename = f"presence_classe_{filiere.nom}_{export_type}.xlsx"

    elif scope == "teacher":
        enseignant_id = request.query_params.get("enseignant_id")
        if not enseignant_id:
            return Response({"message": "ID enseignant requis."}, status=status.HTTP_400_BAD_REQUEST)
        enseignant = Enseignant.objects.select_related("user").filter(id=enseignant_id).first()
        if not enseignant:
            return Response({"message": "Enseignant introuvable."}, status=status.HTTP_404_NOT_FOUND)

        seances = Seance.objects.filter(cours__enseignant=enseignant, qrcode__isnull=False).select_related(
            "cours", "cours__module", "cours__module__filiere"
        )
        temp_seances = TemporarySeance.objects.filter(enseignant=enseignant).select_related(
            "module", "module__filiere"
        )

        if semestre_filter:
            seances = seances.filter(cours__module__semestre=semestre_filter)
            temp_seances = temp_seances.filter(module__semestre=semestre_filter)

        eligible_by_filiere = {item["filiere_id"]: item["total"] for item in Etudiant.objects.values("filiere_id").annotate(total=Count("id"))}
        
        presence_counts = {str(item["seance_id"]): item["total"] for item in Presence.objects.filter(seance_id__in=seances).values("seance_id").annotate(total=Count("id"))}
        temp_presence_counts = {item["seance_id"]: item["total"] for item in TemporaryPresence.objects.filter(seance_id__in=temp_seances).values("seance_id").annotate(total=Count("id"))}

        sessions_list = []
        for s in seances:
            eligible = eligible_by_filiere.get(s.cours.module.filiere_id, 0)
            p_count = presence_counts.get(str(s.id), 0)
            sessions_list.append({
                "date": s.date_seance,
                "module": s.cours.module.nom,
                "filiere": s.cours.module.filiere.nom,
                "semestre": s.cours.module.semestre,
                "type": "Régulier",
                "salle": s.cours.salle or "Non précisée",
                "eligible": eligible,
                "present": p_count,
                "absent": max(eligible - p_count, 0)
            })

        for ts in temp_seances:
            eligible = eligible_by_filiere.get(ts.module.filiere_id, 0)
            p_count = temp_presence_counts.get(ts.id, 0)
            sessions_list.append({
                "date": ts.date_seance,
                "module": ts.module.nom,
                "filiere": ts.module.filiere.nom,
                "semestre": ts.module.semestre,
                "type": "Temporaire",
                "salle": ts.salle or "Non précisée",
                "eligible": eligible,
                "present": p_count,
                "absent": max(eligible - p_count, 0)
            })

        sessions_list.sort(key=lambda x: x["date"], reverse=True)

        details = []
        for s in seances:
            filiere_id = s.cours.module.filiere_id
            stds = Etudiant.objects.filter(filiere_id=filiere_id).select_related("user")
            presences = {p.etudiant_id: p for p in Presence.objects.filter(seance=s)}
            for std in stds:
                p = presences.get(std.id)
                stat = p.statut if p else "absent"
                if export_type == "present" and stat == "absent":
                    continue
                if export_type == "absent" and stat in ["present", "retard"]:
                    continue
                val_time = p.heure_validation if (p and p.heure_validation) else None
                details.append({
                    "date": s.date_seance,
                    "module": s.cours.module.nom,
                    "filiere": s.cours.module.filiere.nom,
                    "semestre": s.cours.module.semestre,
                    "student_name": f"{std.user.prenom} {std.user.nom}",
                    "code_massar": std.code_massar,
                    "status": stat,
                    "validated_at": val_time
                })

        for ts in temp_seances:
            filiere_id = ts.module.filiere_id
            stds = Etudiant.objects.filter(filiere_id=filiere_id).select_related("user")
            presences = {p.etudiant_id: p for p in TemporaryPresence.objects.filter(seance=ts)}
            for std in stds:
                p = presences.get(std.id)
                stat = p.statut if p else "absent"
                if export_type == "present" and stat == "absent":
                    continue
                if export_type == "absent" and stat in ["present", "retard"]:
                    continue
                val_time = p.heure_validation if (p and p.heure_validation) else None
                details.append({
                    "date": ts.date_seance,
                    "module": ts.module.nom,
                    "filiere": ts.module.filiere.nom,
                    "semestre": ts.module.semestre,
                    "student_name": f"{std.user.prenom} {std.user.nom}",
                    "code_massar": std.code_massar,
                    "status": stat,
                    "validated_at": val_time
                })

        details.sort(key=lambda d: d["date"], reverse=True)

        ws1 = wb.active
        ws1.title = "Séances Enseignant"
        apply_common_styles(ws1, "SÉANCES DE L'ENSEIGNANT", f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        ws1["B5"] = f"Enseignant : {enseignant.user.prenom} {enseignant.user.nom}"
        ws1["B5"].font = Font(name="Arial", size=9, bold=True)
        ws1["E5"] = f"Semestre : {semestre_filter or 'Tous les semestres'}"
        ws1["E5"].font = Font(name="Arial", size=9, bold=True)

        headers1 = ["N°", "Date", "Module", "Filière", "Semestre", "Type", "Salle", "Éligibles", "Présents", "Absents", "Taux de Présence"]
        write_table_headers(ws1, 8, headers1)

        current_row = 9
        for idx, row in enumerate(sessions_list, 1):
            ws1.cell(row=current_row, column=2, value=idx)
            ws1.cell(row=current_row, column=3, value=row["date"].strftime("%d/%m/%Y") if isinstance(row["date"], datetime) or hasattr(row["date"], "strftime") else str(row["date"]))
            ws1.cell(row=current_row, column=4, value=row["module"])
            ws1.cell(row=current_row, column=5, value=row["filiere"])
            ws1.cell(row=current_row, column=6, value=row["semestre"])
            ws1.cell(row=current_row, column=7, value=row["type"])
            ws1.cell(row=current_row, column=8, value=row["salle"])
            ws1.cell(row=current_row, column=9, value=row["eligible"])
            ws1.cell(row=current_row, column=10, value=row["present"])
            ws1.cell(row=current_row, column=11, value=row["absent"])
            rate_percentage = round((row["present"] / row["eligible"]) * 100) if row["eligible"] else 0
            ws1.cell(row=current_row, column=12, value=f"{rate_percentage}%")
            current_row += 1

        style_table_rows(ws1, 9, len(sessions_list), center_cols=[2, 3, 6, 7, 8, 9, 10, 11, 12])
        auto_fit_columns(ws1, start_row=8)

        ws2 = wb.create_sheet(title="Détail Émargements")
        apply_common_styles(ws2, "DÉTAIL DES ÉMARGEMENTS ENSEIGNANT", f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        ws2["B5"] = f"Enseignant : {enseignant.user.prenom} {enseignant.user.nom}"
        ws2["B5"].font = Font(name="Arial", size=9, bold=True)

        headers2 = ["N°", "Date", "Module", "Filière", "Semestre", "Nom", "Prénom", "Code Massar", "Statut", "Heure de Validation"]
        write_table_headers(ws2, 8, headers2)

        current_row = 9
        for idx, r in enumerate(details, 1):
            ws2.cell(row=current_row, column=2, value=idx)
            ws2.cell(row=current_row, column=3, value=r["date"].strftime("%d/%m/%Y") if isinstance(r["date"], datetime) or hasattr(r["date"], "strftime") else str(r["date"]))
            ws2.cell(row=current_row, column=4, value=r["module"])
            ws2.cell(row=current_row, column=5, value=r["filiere"])
            ws2.cell(row=current_row, column=6, value=r["semestre"])
            names = r["student_name"].split(" ")
            first_name = names[0]
            last_name = " ".join(names[1:]) if len(names) > 1 else ""
            ws2.cell(row=current_row, column=7, value=last_name)
            ws2.cell(row=current_row, column=8, value=first_name)
            ws2.cell(row=current_row, column=9, value=r["code_massar"])
            ws2.cell(row=current_row, column=10, value="Présent" if r["status"] in ["present", "retard"] else "Absent")

            val_time_str = ""
            if r["validated_at"]:
                try:
                    val_time_str = timezone.localtime(r["validated_at"]).strftime("%d/%m/%Y %H:%M:%S")
                except Exception:
                    val_time_str = str(r["validated_at"])
            ws2.cell(row=current_row, column=11, value=val_time_str)
            current_row += 1

        style_table_rows(ws2, 9, len(details), status_col_idx=10, center_cols=[2, 3, 6, 9, 10, 11])
        auto_fit_columns(ws2, start_row=8)
        ws2.column_dimensions["K"].width = 24

        filename = f"presence_enseignant_{enseignant.user.nom}_{export_type}.xlsx"

    else:
        return Response({"message": "Scope invalide."}, status=status.HTTP_400_BAD_REQUEST)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
